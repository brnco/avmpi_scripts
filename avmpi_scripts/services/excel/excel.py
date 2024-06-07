'''
handles all things Excel / xlsx for AVMPI
'''
import logging
import pathlib
import json
import openpyxl
from pprint import pformat
from openpyxl.utils import get_column_letter


logger = logging.getLogger('main_logger')


def config():
    '''
    loads config from config file located in same directory
    '''
    parent_dirpath = pathlib.Path(__file__).parent.absolute()
    excel_conf_filepath = parent_dirpath / "excel_config.json"
    with open(excel_conf_filepath, "r") as excel_conf_file:
        excel_conf = json.load(excel_conf_file)
    return excel_conf


def load_field_mappings():
    '''
    loads field mappings from config
    '''
    module_dirpath = pathlib.Path(__file__).parent.parent.parent.absolute()
    field_mappings_path = module_dirpath / "field_mappings.json"
    with open(field_mappings_path, "r") as field_mappings_file:
        field_mapping = json.load(field_mappings_file)
    return field_mapping


def get_workbook_type(workbook):
    '''
    looks through the sheets in teh workbook to determine if this is
    Unit metadata or Vendor metadata
    '''
    conf = config()
    unit_sheetnames = list(conf['Unit-AssetsMetadata'].keys())
    vendor_sheetnames = list(conf['Vendor-AssetsMetadata'].keys())
    workbook_sheetnames = [sheet for sheet in workbook.sheetnames]
    if unit_sheetnames == workbook_sheetnames:
        return 'Unit-AssetsMetadata'
    elif vendor_sheetnames == workbook_sheetnames:
        return 'Vendor-AssetsMetadata'
    else:
        return None


def load_all_worksheets(filepath):
    '''
    loads all worksheets from xlsx at filepath into dictionary
    where each key is the worksheet name
    and each value is a list of rows, which themselves are dictionaries
    of key: value paires where the key is the column letter and value is cell value
    '''
    logger.debug(f"loading Excel spreadsheet from {filepath}...")
    workbook = openpyxl.load_workbook(filepath)
    workbook_type = get_workbook_type(workbook)
    if not workbook_type:
        raise RuntimeError("could not identify if workbook is Vendor or Unit type")
    conf = config()
    sheets_data = {}
    for sheet_name in workbook.sheetnames:
        try:
            if conf[workbook_type][sheet_name]['skip']:
                continue
        except KeyError:
            pass
        sheet = workbook[sheet_name]
        start_row = conf[workbook_type][sheet_name]["first_row_with_data"]
        rows_dict = {}
        for row_index, row in enumerate(sheet.iter_rows(values_only=True, min_row=start_row), start=start_row):
            row_data = {}
            for column_index, cell_value in enumerate(row, start=1):
                column_letter = get_column_letter(column_index)
                row_data[column_letter] = cell_value
            if any(cell is not None for cell in row_data.values()):
                rows_dict[row_index] = row_data
        sheets_data[sheet_name] = rows_dict
    return sheets_data


def validate_row(row, required_columns):
    '''
    validates an individual row against its required columns
    '''
    missing_values = []
    for req_col in required_columns:
        if not row[req_col]:
            missing_values.append(req_col)
    return missing_values


def validate_required_fields(rows, record_type):
    '''
    uses rules in field_mappings.json to validate sheet_dict
    '''
    logger.debug("validating worksheet...")
    _field_map = load_field_mappings()
    field_map = _field_map[record_type]
    required_columns = []
    missing_values = []
    for field, mapping in field_map.items():
        try:
            assert mapping['xlsx']['required']
        except (KeyError, TypeError):
            continue
        if mapping['xlsx']['required']:
                required_columns.append(mapping['xlsx']['column'])
    for row in rows:
        missing_columns = validate_row(row, required_columns)
        if missing_columns:
            missing_values.append({"row": rows.index(row), "columns": missing_columns})
    return missing_values
