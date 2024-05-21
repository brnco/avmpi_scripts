'''
handles all things Excel / xlsx for AVMPI
'''
import logging
import openpyxl
from openpyxl.utils import get_column_letter


logger = logging.getLogger('main_logger')


def load_all_worksheets(filepath):
    '''
    loads all worksheets from xlsx at filepath into dictionary
    where each key is the worksheet name
    and each value is a list of rows from that sheet
    '''
    logger.debug(f"loading Excel spreadsheet from {filepath}...")
    workbook = openpyxl.load_workbook(filepath)
    sheets_data = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
        sheets_data[sheet_name] = rows
    return sheets_data


def load_all_worksheets_dict(filepath):
    '''
    loads all worksheets from xlsx at filepath into dictionary
    where each key is the worksheet name
    and each value is a list of rows, which themselves are dictionaries
    of key: value paires where the key is the column letter and value is cell value
    '''
    logger.debug(f"loading Excel spreadsheet from {filepath}...")
    workbook = openpyxl.load_workbook(filepath)
    sheets_data = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows_dict = {}
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_data = {}
            for column_index, cell_value in enumerate(row, start=1):
                column_letter = get_column_letter(column_index)
                row_data[column_letter] = cell_value
            rows_dict[row_index] = row_data
        sheets_data[sheet_name] = rows_dict
    return sheets_data
